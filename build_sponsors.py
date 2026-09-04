# -*- coding: utf-8 -*-
"""
후원처 엑셀(후원처_지도관리.xlsx)  ->  sponsors.json

담당자는 상호명·업종·주소·전화만 입력한다. 좌표(위도·경도)는 비워둬도 된다.
이 스크립트가 카카오 지도로 주소를 좌표로 바꿔 자동으로 채운다.
(원래 구글시트 Apps Script 'Untitled1.js' 가 하던 일을 파이썬으로 옮긴 것)

카카오 REST 키는 코드에 두지 않는다. GitHub Actions Secret(KAKAO_REST_KEY)에서 꺼내 쓴다.
한 번 찾은 좌표는 _coords_cache.json 에 저장해, 다음부터는 다시 검색하지 않는다.
"""
import json, os, re, sys, time, urllib.parse, urllib.request
from pathlib import Path

import openpyxl

XLSX  = Path("북원_후원업체_지도관리.xlsx")
OUT   = Path("sponsors.json")
CACHE = Path("_coords_cache.json")
KEY   = os.environ.get("KAKAO_REST_KEY", "").strip()

errors = []
def err(row, msg): errors.append(f"{row}행: {msg}")

def s(v):
    return "" if v is None else str(v).strip()

# 주민등록번호만 차단한다. (전화는 후원 가게 번호라 010 이어도 정상)
JUMIN = re.compile(r"\b\d{6}[-]\d{7}\b")

def load_cache():
    if CACHE.exists():
        try: return json.loads(CACHE.read_text(encoding="utf-8"))
        except Exception: return {}
    return {}

def kakao_geocode(addr):
    """주소 → (lat, lng). 주소검색 실패 시 키워드검색. 못 찾으면 None."""
    if not KEY:
        raise RuntimeError("KAKAO_REST_KEY 가 없습니다. GitHub Actions Secret 에 등록했는지 확인하세요.")
    for kind in ("address", "keyword"):
        url = (f"https://dapi.kakao.com/v2/local/search/{kind}.json?query="
               + urllib.parse.quote(addr))
        req = urllib.request.Request(url, headers={"Authorization": "KakaoAK " + KEY})
        try:
            with urllib.request.urlopen(req, timeout=10) as r:
                docs = json.loads(r.read().decode("utf-8")).get("documents", [])
        except Exception as e:
            raise RuntimeError(f"카카오 API 호출 실패: {e}")
        if docs:
            return float(docs[0]["y"]), float(docs[0]["x"])
        time.sleep(0.1)
    return None

def main():
    if not XLSX.exists():
        print(f"::error::{XLSX} 파일이 없습니다.")
        sys.exit(1)

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    if "후원처" not in wb.sheetnames:
        print("::error::'후원처' 시트가 없습니다. 시트 이름을 바꾸지 마세요.")
        sys.exit(1)
    ws = wb["후원처"]

    head = [s(c.value) for c in ws[4]]
    need = ["게시","상호명","업종","주소","전화","위도","경도","비고"]
    for h in need:
        if h not in head:
            errors.append(f"'후원처' 시트 4행에 '{h}' 열이 없습니다. 열 이름을 바꾸지 마세요.")
    if errors:
        for e in errors: print("  ✗ " + e)
        print(f"::error::엑셀 열 이름 문제 {len(errors)}건")
        sys.exit(1)
    ix = {h: head.index(h) for h in need}

    cache = load_cache()
    rows_out = []
    to_geocode = []   # (엑셀행, 주소)
    staged = []       # 임시 저장

    for r in range(5, ws.max_row + 1):
        row = [c.value for c in ws[r]]
        if not any(s(v) for v in row): continue

        pub = s(row[ix["게시"]]).upper()
        if pub in ("X", "×"): continue
        if pub != "O":
            err(r, f"게시 칸은 O 또는 X 만 (지금: '{s(row[ix['게시']])}')"); continue

        name = s(row[ix["상호명"]])
        addr = s(row[ix["주소"]])
        if not name: err(r, "상호명이 비어 있습니다"); continue
        if not addr: err(r, f"'{name}' 의 주소가 비어 있습니다"); continue
        if JUMIN.search(" ".join(s(v) for v in row)):
            err(r, "주민등록번호로 보이는 값이 있습니다. 공개 저장소라 올릴 수 없습니다")

        cat  = s(row[ix["업종"]])
        tel  = s(row[ix["전화"]])
        memo = s(row[ix["비고"]])
        lat  = s(row[ix["위도"]])
        lng  = s(row[ix["경도"]])

        rec = {"name": name, "cat": cat, "addr": addr, "tel": tel, "memo": memo,
               "lat": lat, "lng": lng, "row": r}

        if lat and lng:
            pass                       # 엑셀에 좌표를 직접 넣은 경우 그대로
        elif addr in cache:
            rec["lat"], rec["lng"] = cache[addr]   # 전에 찾아둔 좌표
        else:
            to_geocode.append((r, addr))
        staged.append(rec)

    if errors:
        print("\n엑셀에서 고쳐야 할 곳이 있습니다. 지도는 그대로 둡니다.\n")
        for e in errors: print("  ✗ " + e)
        print(f"::error::엑셀 오류 {len(errors)}건")
        sys.exit(1)

    # 새 주소만 좌표 검색
    fails = []
    for r, addr in to_geocode:
        try:
            coord = kakao_geocode(addr)
        except RuntimeError as e:
            print("\n" + str(e))
            print("::error::" + str(e))
            sys.exit(1)
        if coord:
            cache[addr] = [coord[0], coord[1]]
        else:
            fails.append(f"{r}행: 주소를 지도에서 못 찾음 → {addr}")
        time.sleep(0.12)

    # 캐시 반영
    for rec in staged:
        if (not rec["lat"] or not rec["lng"]) and rec["addr"] in cache:
            rec["lat"], rec["lng"] = cache[rec["addr"]]

    if fails:
        print("\n좌표를 못 찾은 주소가 있습니다. 지도는 그대로 둡니다.\n")
        for f in fails: print("  ✗ " + f)
        print("주소를 정확히 고치거나, 위도·경도를 직접 넣어 주세요.")
        print(f"::error::좌표 검색 실패 {len(fails)}건")
        sys.exit(1)

    # index.html 이 쓰는 배열 형태로 출력: [연번, 상호명, 업종, 주소, 전화, 위도, 경도, 비고]
    header = ["연번","상호명","업종","주소","전화","위도","경도","비고"]
    arr = [header]
    for i, rec in enumerate(staged, start=1):
        arr.append([i, rec["name"], rec["cat"], rec["addr"], rec["tel"],
                    rec["lat"], rec["lng"], rec["memo"]])

    OUT.write_text(json.dumps(arr, ensure_ascii=False, indent=1), encoding="utf-8")
    CACHE.write_text(json.dumps(cache, ensure_ascii=False, indent=1), encoding="utf-8")
    print(f"✓ {OUT} 생성 — 후원처 {len(staged)}곳 "
          f"(새로 좌표 찾음 {len(to_geocode)}곳, 캐시 {len(cache)}건)")

if __name__ == "__main__":
    main()
